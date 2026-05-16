"""
T175: Formatted Excel exports.
"""
from __future__ import annotations

import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from analytics_engine.reports.disclaimers import REPORT_DISCLAIMER


def export_holdings_xlsx(holdings: list[dict], filename: str | None = None) -> str:
    """Export holdings data to a formatted Excel file. Returns path."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Holdings"

    headers = ["Symbol", "Exchange", "Qty", "Avg Price", "Last Price", "P&L", "Sector"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row, h in enumerate(holdings, 2):
        ws.cell(row=row, column=1, value=h.get("symbol", h.get("tradingsymbol", "")))
        ws.cell(row=row, column=2, value=h.get("exchange", ""))
        ws.cell(row=row, column=3, value=h.get("quantity", 0))
        ws.cell(row=row, column=4, value=float(h.get("average_price", 0)))
        ws.cell(row=row, column=5, value=float(h.get("last_price", 0)))
        ws.cell(row=row, column=6, value=float(h.get("pnl", 0)))
        ws.cell(row=row, column=7, value=h.get("sector", ""))

    # Disclaimer row
    disc_row = len(holdings) + 3
    ws.cell(row=disc_row, column=1, value=REPORT_DISCLAIMER)
    ws.cell(row=disc_row, column=1).font = Font(italic=True, size=9)

    if not filename:
        tmpdir = tempfile.mkdtemp()
        filename = str(Path(tmpdir) / "holdings_export.xlsx")

    wb.save(filename)
    return filename
